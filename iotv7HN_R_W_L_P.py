import openpyxl
import time 
import math
import matplotlib.pyplot as plt
import array as arr

def main():
    #book = openpyxl.load_workbook('C:\Users\shahab\OneDrive\FA_paper\py\Sample.xlsx')
    book = openpyxl.load_workbook('temp_mean_105_days_actual_predictions.xlsx')
    #book = openpyxl.Workbook()
   # book.create_sheet('Sample')


    #sheet = book["Sample"]
    sheet = book.active



    NUM = 2612
    MAX=80
    MIN=-46
    MINs=MIN
    ri=0
    ci=0
    NMM=105
    SStaticBit=395

    DDynamicBit=70




    if MINs<0:
      MIN=MIN- MINs
      MAX=MAX- MINs
      for R in range(NUM):
         sheet.cell(row= R+1+ri, column=1+ci).value = sheet.cell(row= R+1+ri, column=1+ci).value - MINs
         sheet.cell(row= R+1+ri, column=2+ci).value = sheet.cell(row= R+1+ri, column=2+ci).value - MINs

#ijad magadir beyn max v min , v barbar sefr garar dadan un ha
    for R in range(MIN,MAX+1):
       sheet.cell(row= R-MIN+1+ri, column=3+ci).value = R
       sheet.cell(row= R-MIN+1+ri, column=4+ci).value = 0

#shomaresh magadir
    for R in range(NMM):
          sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value = sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value+1


#copy
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value

    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value



#sorting
#     for T in range(MAX-MIN+1):
#        for R in range(((MAX-MIN+1)-1) - T):

#           if sheet.cell(row= R+2+ri, column=6+ci).value > sheet.cell(row= R+1+ri, column=6+ci).value:
#             temp = sheet.cell(row= R+2+ri, column=6+ci).value
#             temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
#             sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
#             sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
#             sheet.cell(row= R+1+ri, column=6+ci).value = temp
#             sheet.cell(row= R+1+ri, column=7+ci).value = temp1  


# #sorting
#     for T in range(MAX-MIN+1):
#        for R in range(((MAX-MIN+1)-1) - T):

#           if sheet.cell(row= R+2+ri, column=7+ci).value > sheet.cell(row= R+1+ri, column=7+ci).value:
#             temp = sheet.cell(row= R+2+ri, column=6+ci).value
#             temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
#             sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
#             sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
#             sheet.cell(row= R+1+ri, column=6+ci).value = temp
#             sheet.cell(row= R+1+ri, column=7+ci).value = temp1        


#mohasebe
#copy
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=9+ci).value = str(sheet.cell(row= R+1+ri, column=6+ci).value)
       sheet.cell(row= R+1+ri, column=12+ci).value = str(sheet.cell(row= R+1+ri, column=6+ci).value)

       sheet.cell(row= R+1+ri, column=10+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
       sheet.cell(row= R+1+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value

       sheet.cell(row= R+1+ri, column=14+ci).value = '*'
       sheet.cell(row= R+1+ri, column=15+ci).value = '*'

    # for k in range(MAX-MIN):
    #    R=(MAX-MIN+1)-(k+1)

    #    STe=sheet.cell(row= R+1+ri, column=9+ci).value + sheet.cell(row= R+ri, column=9+ci).value
    #    VTe=sheet.cell(row= R+1+ri, column=10+ci).value + sheet.cell(row= R+ri, column=10+ci).value

    #    sheet.cell(row= k+1+ MAX-MIN+1+ri, column=12+ci).value = STe
    #    sheet.cell(row= k+1+ MAX-MIN+1+ri, column=13+ci).value = VTe
    #    sheet.cell(row= k+1+ MAX-MIN+1+ri, column=14+ci).value = sheet.cell(row= R+1+ri, column=9+ci).value
    #    sheet.cell(row= k+1+ MAX-MIN+1+ri, column=15+ci).value = sheet.cell(row= R+ri, column=9+ci).value

    #    E=R
    #    while E>=1:
    #     if E==1:
    #       B=E
    #       WM=MAX-MIN+1
    #       while B< WM:
    #         sheet.cell(row= WM+ri, column=9+ci).value=sheet.cell(row= WM+ri-1, column=9+ci).value
    #         sheet.cell(row= WM+ri, column=10+ci).value=sheet.cell(row= WM+ri-1, column=10+ci).value

    #         WM=WM-1

    #       sheet.cell(row= E+ri, column=9+ci).value = STe
    #       sheet.cell(row= E+ri, column=10+ci).value = VTe
    #       break
          
    #     else:
    #        if VTe<= (sheet.cell(row= E+ri-1, column=10+ci).value):
    #         B=E
    #         WM=MAX-MIN+1

    #         while B< WM:
    #           sheet.cell(row= WM+ri, column=9+ci).value=sheet.cell(row= WM+ri-1, column=9+ci).value
    #           sheet.cell(row= WM+ri, column=10+ci).value=sheet.cell(row= WM+ri-1, column=10+ci).value

    #           WM=WM-1

    #         sheet.cell(row= E+ri, column=9+ci).value = STe
    #         sheet.cell(row= E+ri, column=10+ci).value = VTe
    #         break

    #     E=E-1

    # for R in range(MAX-MIN+1):
    #    RTe=sheet.cell(row= R+1+ri, column=12+ci).value
    #    coun=0
    #    End=MAX-MIN+1+MAX-MIN
    #    FSTe=""
    #    while RTe != sheet.cell(row= 1+ri, column=9+ci).value:
    #      Ind=MAX-MIN+1
    #      while Ind< End :
    #       if sheet.cell(row= Ind+1+ri, column=14+ci).value==RTe:
    #         FSTe=FSTe+"0"
    #         coun=coun+1
    #         RTe=sheet.cell(row= Ind+1+ri, column=12+ci).value
    #         break

    #       if sheet.cell(row= Ind+1+ri, column=15+ci).value==RTe:
    #         FSTe=FSTe+"1"
    #         coun=coun+1
    #         RTe=sheet.cell(row= Ind+1+ri, column=12+ci).value
    #         break
    #       Ind=Ind+1
    #    sheet.cell(row= R+1+ri, column=17+ci).value =  FSTe
    #    sheet.cell(row= R+1+ri, column=18+ci).value =  coun

    for R in range(MIN,MAX+1):
      sheet.cell(row= R-MIN+1+ri, column=3+ci).value = R
      sheet.cell(row= R-MIN+1+ri, column=4+ci).value = 0

      sheet.cell(row= R-MIN+1+ri, column=23+ci).value = R
      sheet.cell(row= R-MIN+1+ri, column=24+ci).value = 0

         
                  

#shomaresh magadir
    p=0
    pl=0
    for R in range(NMM,NUM):
          sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value = sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value+1
          sheet.cell(row= (sheet.cell(row= R+1+ri, column=2+ci).value)-MIN+1+ri, column=24+ci).value = sheet.cell(row= (sheet.cell(row= R+1+ri, column=2+ci).value)-MIN+1+ri, column=24+ci).value+1
          if ((R+1)%NMM)==0:
#########################################################3


            # #copy
            #     for R in range(MAX-MIN+1):
            #        sheet.cell(row= R+1+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value

            #     for R in range(MAX-MIN+1):
            #        sheet.cell(row= R+1+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value

                print(round((R/NUM)*100), end="\r")
            ###############################################################################################copy
                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=20+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=21+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value


            ###################################################################################################copy learning
                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=26+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=27+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value





            ######################################################################################################sorting
                for T in range(MAX-MIN+1):
                   for R in range(((MAX-MIN+1)-1) - T):

                      if sheet.cell(row= R+2+ri, column=6+ci).value > sheet.cell(row= R+1+ri, column=6+ci).value:
                        temp = sheet.cell(row= R+2+ri, column=6+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
                        sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
                        sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
                        sheet.cell(row= R+1+ri, column=6+ci).value = temp
                        sheet.cell(row= R+1+ri, column=7+ci).value = temp1  


                        temp = sheet.cell(row= R+2+ri, column=20+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=21+ci).value
                        sheet.cell(row= R+2+ri, column=20+ci).value = sheet.cell(row= R+1+ri, column=20+ci).value
                        sheet.cell(row= R+2+ri, column=21+ci).value = sheet.cell(row= R+1+ri, column=21+ci).value
                        sheet.cell(row= R+1+ri, column=20+ci).value = temp
                        sheet.cell(row= R+1+ri, column=21+ci).value = temp1  




            #sorting
                for T in range(MAX-MIN+1):
                   for R in range(((MAX-MIN+1)-1) - T):

                      if sheet.cell(row= R+2+ri, column=7+ci).value > sheet.cell(row= R+1+ri, column=7+ci).value:
                        temp = sheet.cell(row= R+2+ri, column=6+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
                        sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
                        sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
                        sheet.cell(row= R+1+ri, column=6+ci).value = temp
                        sheet.cell(row= R+1+ri, column=7+ci).value = temp1        



                        temp = sheet.cell(row= R+2+ri, column=20+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=21+ci).value
                        sheet.cell(row= R+2+ri, column=20+ci).value = sheet.cell(row= R+1+ri, column=20+ci).value
                        sheet.cell(row= R+2+ri, column=21+ci).value = sheet.cell(row= R+1+ri, column=21+ci).value
                        sheet.cell(row= R+1+ri, column=20+ci).value = temp
                        sheet.cell(row= R+1+ri, column=21+ci).value = temp1  

###########################################################################Final Sorting
                pointer= sheet.cell(row= 1+ri, column=6+ci).value
                er=0
                T=0
                while(T<MAX-MIN+1):
                  
                  #print("er"+str(er))
                  #print("T"+str(T))

                  if sheet.cell(row= T+1+ri, column=7+ci).value == 0:
                    er=er+1
                    for k in range(MAX-MIN+1):
                      if pointer + er<= MAX :
                        if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer +er ):
                          temp = sheet.cell(row= T+1+ri, column=6+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
                          sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
                          sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
                          sheet.cell(row= k+1+ri, column=6+ci).value = temp
                          sheet.cell(row= k+1+ri, column=7+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=20+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=21+ci).value
                          sheet.cell(row= T+1+ri, column=20+ci).value = sheet.cell(row= k+1+ri, column=20+ci).value
                          sheet.cell(row= T+1+ri, column=21+ci).value = sheet.cell(row= k+1+ri, column=21+ci).value
                          sheet.cell(row= k+1+ri, column=20+ci).value = temp
                          sheet.cell(row= k+1+ri, column=21+ci).value = temp1
                          T=T+1

                      if pointer - er>= MIN :
                        if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer -er ):
                          temp = sheet.cell(row= T+1+ri, column=6+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
                          sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
                          sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
                          sheet.cell(row= k+1+ri, column=6+ci).value = temp
                          sheet.cell(row= k+1+ri, column=7+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=20+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=21+ci).value
                          sheet.cell(row= T+1+ri, column=20+ci).value = sheet.cell(row= k+1+ri, column=20+ci).value
                          sheet.cell(row= T+1+ri, column=21+ci).value = sheet.cell(row= k+1+ri, column=21+ci).value
                          sheet.cell(row= k+1+ri, column=20+ci).value = temp
                          sheet.cell(row= k+1+ri, column=21+ci).value = temp1
                          T=T+1
                  
                  else:
                    #print("else")
                    T=T+1


                #print("While End")




            ######################################################################################################sorting learning 23 24 ..... 26 27
                for T in range(MAX-MIN+1):
                   for R in range(((MAX-MIN+1)-1) - T):

                      if sheet.cell(row= R+2+ri, column=23+ci).value > sheet.cell(row= R+1+ri, column=23+ci).value:
                        temp = sheet.cell(row= R+2+ri, column=23+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=24+ci).value
                        sheet.cell(row= R+2+ri, column=23+ci).value = sheet.cell(row= R+1+ri, column=23+ci).value
                        sheet.cell(row= R+2+ri, column=24+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value
                        sheet.cell(row= R+1+ri, column=23+ci).value = temp
                        sheet.cell(row= R+1+ri, column=24+ci).value = temp1  


                        temp = sheet.cell(row= R+2+ri, column=26+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=27+ci).value
                        sheet.cell(row= R+2+ri, column=26+ci).value = sheet.cell(row= R+1+ri, column=26+ci).value
                        sheet.cell(row= R+2+ri, column=27+ci).value = sheet.cell(row= R+1+ri, column=27+ci).value
                        sheet.cell(row= R+1+ri, column=26+ci).value = temp
                        sheet.cell(row= R+1+ri, column=27+ci).value = temp1  




            #sorting
                for T in range(MAX-MIN+1):
                   for R in range(((MAX-MIN+1)-1) - T):

                      if sheet.cell(row= R+2+ri, column=24+ci).value > sheet.cell(row= R+1+ri, column=24+ci).value:
                        temp = sheet.cell(row= R+2+ri, column=23+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=24+ci).value
                        sheet.cell(row= R+2+ri, column=23+ci).value = sheet.cell(row= R+1+ri, column=23+ci).value
                        sheet.cell(row= R+2+ri, column=24+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value
                        sheet.cell(row= R+1+ri, column=23+ci).value = temp
                        sheet.cell(row= R+1+ri, column=24+ci).value = temp1        



                        temp = sheet.cell(row= R+2+ri, column=26+ci).value
                        temp1 = sheet.cell(row= R+2+ri, column=27+ci).value
                        sheet.cell(row= R+2+ri, column=26+ci).value = sheet.cell(row= R+1+ri, column=26+ci).value
                        sheet.cell(row= R+2+ri, column=27+ci).value = sheet.cell(row= R+1+ri, column=27+ci).value
                        sheet.cell(row= R+1+ri, column=26+ci).value = temp
                        sheet.cell(row= R+1+ri, column=27+ci).value = temp1  

#############################################################################################################Final Sorting learning 23 24 ..... 26 27
                pointer= sheet.cell(row= 1+ri, column=23+ci).value
                er=0
                T=0
                while(T<MAX-MIN+1):
                  
                  #print("er"+str(er))
                  #print("T"+str(T))

                  if sheet.cell(row= T+1+ri, column=24+ci).value == 0:
                    er=er+1
                    for k in range(MAX-MIN+1):
                      if pointer + er<= MAX :
                        if (sheet.cell(row= k+1+ri, column=24+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=23+ci).value == pointer +er ):
                          temp = sheet.cell(row= T+1+ri, column=23+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=24+ci).value
                          sheet.cell(row= T+1+ri, column=23+ci).value = sheet.cell(row= k+1+ri, column=23+ci).value
                          sheet.cell(row= T+1+ri, column=24+ci).value = sheet.cell(row= k+1+ri, column=24+ci).value
                          sheet.cell(row= k+1+ri, column=23+ci).value = temp
                          sheet.cell(row= k+1+ri, column=24+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=26+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=27+ci).value
                          sheet.cell(row= T+1+ri, column=26+ci).value = sheet.cell(row= k+1+ri, column=26+ci).value
                          sheet.cell(row= T+1+ri, column=27+ci).value = sheet.cell(row= k+1+ri, column=27+ci).value
                          sheet.cell(row= k+1+ri, column=26+ci).value = temp
                          sheet.cell(row= k+1+ri, column=27+ci).value = temp1
                          T=T+1

                      if pointer - er>= MIN :
                        if (sheet.cell(row= k+1+ri, column=24+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=23+ci).value == pointer -er ):
                          temp = sheet.cell(row= T+1+ri, column=23+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=24+ci).value
                          sheet.cell(row= T+1+ri, column=23+ci).value = sheet.cell(row= k+1+ri, column=23+ci).value
                          sheet.cell(row= T+1+ri, column=24+ci).value = sheet.cell(row= k+1+ri, column=24+ci).value
                          sheet.cell(row= k+1+ri, column=23+ci).value = temp
                          sheet.cell(row= k+1+ri, column=24+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=26+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=27+ci).value
                          sheet.cell(row= T+1+ri, column=26+ci).value = sheet.cell(row= k+1+ri, column=26+ci).value
                          sheet.cell(row= T+1+ri, column=27+ci).value = sheet.cell(row= k+1+ri, column=27+ci).value
                          sheet.cell(row= k+1+ri, column=26+ci).value = temp
                          sheet.cell(row= k+1+ri, column=27+ci).value = temp1
                          T=T+1
                  
                  else:
                    #print("else")
                    T=T+1


                #print("While End")
##########################################vazn
               # StaticBit=22
                StaticBit=SStaticBit
               # DynamicBit=4
                DynamicBit=DDynamicBit

                ChangeBit= int((MAX-MIN+1)/DynamicBit)
                for R in range(MIN,MAX+1):

                    if ((R+1)%ChangeBit)==0 and DynamicBit != 1 :
                        DynamicBit= DynamicBit-1

                    if sheet.cell(row= R+1+ri, column=7+ci).value != 0:
                        sheet.cell(row= R+1+ri, column=7+ci).value =sheet.cell(row= R+1+ri, column=7+ci).value + (StaticBit)
                    else: 
                        sheet.cell(row= R+1+ri, column=7+ci).value =sheet.cell(row= R+1+ri, column=7+ci).value + (DynamicBit)

                   
 
##########################################vazn learning
                StaticBit=SStaticBit
                DynamicBit=DDynamicBit
                ChangeBit= int((MAX-MIN+1)/DynamicBit)
                for R in range(MIN,MAX+1):
                    if ((R+1)%ChangeBit)==0 and DynamicBit != 1 :
                        DynamicBit= DynamicBit-1

                    if sheet.cell(row= R+1+ri, column=24+ci).value != 0:
                        sheet.cell(row= R+1+ri, column=24+ci).value =sheet.cell(row= R+1+ri, column=24+ci).value + (StaticBit)
                    else: 
                        sheet.cell(row= R+1+ri, column=24+ci).value =sheet.cell(row= R+1+ri, column=24+ci).value + (DynamicBit)
                        

################################################################################################################## mohasebe
            #


############################################################################# Preparation
            #copy
                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=9+ci).value = str(sheet.cell(row= R+1+ri, column=6+ci).value)
                   sheet.cell(row= R+1+ri, column=12+ci).value = str(sheet.cell(row= R+1+ri, column=6+ci).value)

                   sheet.cell(row= R+1+ri, column=10+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
                   sheet.cell(row= R+1+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value

                   sheet.cell(row= R+1+ri, column=14+ci).value = '*'
                   sheet.cell(row= R+1+ri, column=15+ci).value = '*'
############################################################################# Preparation learning 23 24 ....>> 29 30.. 32 33... 34 35
            #copy
                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=29+ci).value = str(sheet.cell(row= R+1+ri, column=23+ci).value)
                   sheet.cell(row= R+1+ri, column=32+ci).value = str(sheet.cell(row= R+1+ri, column=23+ci).value)

                   sheet.cell(row= R+1+ri, column=30+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value
                   sheet.cell(row= R+1+ri, column=33+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value

                   sheet.cell(row= R+1+ri, column=34+ci).value = '*'
                   sheet.cell(row= R+1+ri, column=35+ci).value = '*'



################################################################################huffman
                for k in range(MAX-MIN):
                   R=(MAX-MIN+1)-(k+1)

                   STe=sheet.cell(row= R+1+ri, column=9+ci).value + sheet.cell(row= R+ri, column=9+ci).value
                   VTe=sheet.cell(row= R+1+ri, column=10+ci).value + sheet.cell(row= R+ri, column=10+ci).value

                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=12+ci).value = STe
                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=13+ci).value = VTe
                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=14+ci).value = sheet.cell(row= R+1+ri, column=9+ci).value
                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=15+ci).value = sheet.cell(row= R+ri, column=9+ci).value

                   E=R
                   while E>=1:
                    if E==1:
                      B=E
                      WM=MAX-MIN+1
                      while B< WM:
                        sheet.cell(row= WM+ri, column=9+ci).value=sheet.cell(row= WM+ri-1, column=9+ci).value
                        sheet.cell(row= WM+ri, column=10+ci).value=sheet.cell(row= WM+ri-1, column=10+ci).value

                        WM=WM-1

                      sheet.cell(row= E+ri, column=9+ci).value = STe
                      sheet.cell(row= E+ri, column=10+ci).value = VTe
                      break
                      
                    else:
                       if VTe<= (sheet.cell(row= E+ri-1, column=10+ci).value):
                        B=E
                        WM=MAX-MIN+1

                        while B< WM:
                          sheet.cell(row= WM+ri, column=9+ci).value=sheet.cell(row= WM+ri-1, column=9+ci).value
                          sheet.cell(row= WM+ri, column=10+ci).value=sheet.cell(row= WM+ri-1, column=10+ci).value

                          WM=WM-1

                        sheet.cell(row= E+ri, column=9+ci).value = STe
                        sheet.cell(row= E+ri, column=10+ci).value = VTe
                        break

                    E=E-1

                for R in range(MAX-MIN+1):
                   RTe=sheet.cell(row= R+1+ri, column=12+ci).value
                   coun=0
                   End=MAX-MIN+1+MAX-MIN
                   FSTe=""
                   while RTe != sheet.cell(row= 1+ri, column=9+ci).value:
                     Ind=MAX-MIN+1
                     while Ind< End :
                      if sheet.cell(row= Ind+1+ri, column=14+ci).value==RTe:
                        FSTe=FSTe+"0"
                        coun=coun+1
                        RTe=sheet.cell(row= Ind+1+ri, column=12+ci).value
                        break

                      if sheet.cell(row= Ind+1+ri, column=15+ci).value==RTe:
                        FSTe=FSTe+"1"
                        coun=coun+1
                        RTe=sheet.cell(row= Ind+1+ri, column=12+ci).value
                        break
                      Ind=Ind+1
                   sheet.cell(row= R+1+ri, column=17+ci).value =  FSTe
                   sheet.cell(row= R+1+ri, column=18+ci).value =  coun


################################################################################huffman learning  29 30.. 32 33... 34 35 .. 37 38
                for k in range(MAX-MIN):
                   R=(MAX-MIN+1)-(k+1)

                   STe=sheet.cell(row= R+1+ri, column=29+ci).value + sheet.cell(row= R+ri, column=29+ci).value
                   VTe=sheet.cell(row= R+1+ri, column=30+ci).value + sheet.cell(row= R+ri, column=30+ci).value

                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=32+ci).value = STe
                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=33+ci).value = VTe
                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=34+ci).value = sheet.cell(row= R+1+ri, column=29+ci).value
                   sheet.cell(row= k+1+ MAX-MIN+1+ri, column=35+ci).value = sheet.cell(row= R+ri, column=29+ci).value

                   E=R
                   while E>=1:
                    if E==1:
                      B=E
                      WM=MAX-MIN+1
                      while B< WM:
                        sheet.cell(row= WM+ri, column=29+ci).value=sheet.cell(row= WM+ri-1, column=29+ci).value
                        sheet.cell(row= WM+ri, column=30+ci).value=sheet.cell(row= WM+ri-1, column=30+ci).value

                        WM=WM-1

                      sheet.cell(row= E+ri, column=29+ci).value = STe
                      sheet.cell(row= E+ri, column=30+ci).value = VTe
                      break
                      
                    else:
                       if VTe<= (sheet.cell(row= E+ri-1, column=30+ci).value):
                        B=E
                        WM=MAX-MIN+1

                        while B< WM:
                          sheet.cell(row= WM+ri, column=29+ci).value=sheet.cell(row= WM+ri-1, column=29+ci).value
                          sheet.cell(row= WM+ri, column=30+ci).value=sheet.cell(row= WM+ri-1, column=30+ci).value

                          WM=WM-1

                        sheet.cell(row= E+ri, column=29+ci).value = STe
                        sheet.cell(row= E+ri, column=30+ci).value = VTe
                        break

                    E=E-1

                for R in range(MAX-MIN+1):
                   RTe=sheet.cell(row= R+1+ri, column=32+ci).value
                   coun=0
                   End=MAX-MIN+1+MAX-MIN
                   FSTe=""
                   while RTe != sheet.cell(row= 1+ri, column=29+ci).value:
                     Ind=MAX-MIN+1
                     while Ind< End :
                      if sheet.cell(row= Ind+1+ri, column=34+ci).value==RTe:
                        FSTe=FSTe+"0"
                        coun=coun+1
                        RTe=sheet.cell(row= Ind+1+ri, column=32+ci).value
                        break

                      if sheet.cell(row= Ind+1+ri, column=35+ci).value==RTe:
                        FSTe=FSTe+"1"
                        coun=coun+1
                        RTe=sheet.cell(row= Ind+1+ri, column=32+ci).value
                        break
                      Ind=Ind+1
                   sheet.cell(row= R+1+ri, column=37+ci).value =  FSTe
                   sheet.cell(row= R+1+ri, column=38+ci).value =  coun









#########################################################################################################
                x=0
                xl=0
                for R in range(MAX-MIN+1):
                  sheet.cell(row= R+1+ri, column=19+ci).value = (sheet.cell(row= R+1+ri, column=18+ci).value)*(sheet.cell(row= R+1+ri, column=21+ci).value)
                  x=x+sheet.cell(row= R+1+ri, column=19+ci).value

                  sheet.cell(row= R+1+ri, column=39+ci).value = (sheet.cell(row= R+1+ri, column=38+ci).value)*(sheet.cell(row= R+1+ri, column=27+ci).value)
                  xl=xl+sheet.cell(row= R+1+ri, column=39+ci).value

                

                p=p+x
                pl=pl+xl


###############################################
            #copy
                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value

                for R in range(MAX-MIN+1):
                   sheet.cell(row= R+1+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value


###########################0000
                for R in range(MIN,MAX+1):
                  sheet.cell(row= R-MIN+1+ri, column=3+ci).value = R
                  sheet.cell(row= R-MIN+1+ri, column=4+ci).value = 0

                for R in range(MIN,MAX+1):
                  sheet.cell(row= R-MIN+1+ri, column=23+ci).value = R
                  sheet.cell(row= R-MIN+1+ri, column=24+ci).value = 0

#############################################################################################################################33








###########################################################################copy
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=20+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=21+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value

    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=26+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=27+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value

###############################################################sorting
    for T in range(MAX-MIN+1):
       for R in range(((MAX-MIN+1)-1) - T):

          if sheet.cell(row= R+2+ri, column=6+ci).value > sheet.cell(row= R+1+ri, column=6+ci).value:
            temp = sheet.cell(row= R+2+ri, column=6+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
            sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
            sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
            sheet.cell(row= R+1+ri, column=6+ci).value = temp
            sheet.cell(row= R+1+ri, column=7+ci).value = temp1  


            temp = sheet.cell(row= R+2+ri, column=20+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=21+ci).value
            sheet.cell(row= R+2+ri, column=20+ci).value = sheet.cell(row= R+1+ri, column=20+ci).value
            sheet.cell(row= R+2+ri, column=21+ci).value = sheet.cell(row= R+1+ri, column=21+ci).value
            sheet.cell(row= R+1+ri, column=20+ci).value = temp
            sheet.cell(row= R+1+ri, column=21+ci).value = temp1  

##############################################################sorting
    for T in range(MAX-MIN+1):
       for R in range(((MAX-MIN+1)-1) - T):

          if sheet.cell(row= R+2+ri, column=7+ci).value > sheet.cell(row= R+1+ri, column=7+ci).value:
            temp = sheet.cell(row= R+2+ri, column=6+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
            sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
            sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
            sheet.cell(row= R+1+ri, column=6+ci).value = temp
            sheet.cell(row= R+1+ri, column=7+ci).value = temp1        



            temp = sheet.cell(row= R+2+ri, column=20+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=21+ci).value
            sheet.cell(row= R+2+ri, column=20+ci).value = sheet.cell(row= R+1+ri, column=20+ci).value
            sheet.cell(row= R+2+ri, column=21+ci).value = sheet.cell(row= R+1+ri, column=21+ci).value
            sheet.cell(row= R+1+ri, column=20+ci).value = temp
            sheet.cell(row= R+1+ri, column=21+ci).value = temp1  

###########################################################################Final Sorting
    pointer= sheet.cell(row= 1+ri, column=6+ci).value
    er=0
    T=0
    while(T<MAX-MIN+1):
      
      #print("er"+str(er))
      #print("T"+str(T))

      if sheet.cell(row= T+1+ri, column=7+ci).value == 0:
        er=er+1
        for k in range(MAX-MIN+1):
          if pointer + er<= MAX :
            if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer +er ):
              temp = sheet.cell(row= T+1+ri, column=6+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
              sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
              sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
              sheet.cell(row= k+1+ri, column=6+ci).value = temp
              sheet.cell(row= k+1+ri, column=7+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=20+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=21+ci).value
              sheet.cell(row= T+1+ri, column=20+ci).value = sheet.cell(row= k+1+ri, column=20+ci).value
              sheet.cell(row= T+1+ri, column=21+ci).value = sheet.cell(row= k+1+ri, column=21+ci).value
              sheet.cell(row= k+1+ri, column=20+ci).value = temp
              sheet.cell(row= k+1+ri, column=21+ci).value = temp1
              T=T+1

          if pointer - er>= MIN :
            if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer -er ):
              temp = sheet.cell(row= T+1+ri, column=6+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
              sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
              sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
              sheet.cell(row= k+1+ri, column=6+ci).value = temp
              sheet.cell(row= k+1+ri, column=7+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=20+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=21+ci).value
              sheet.cell(row= T+1+ri, column=20+ci).value = sheet.cell(row= k+1+ri, column=20+ci).value
              sheet.cell(row= T+1+ri, column=21+ci).value = sheet.cell(row= k+1+ri, column=21+ci).value
              sheet.cell(row= k+1+ri, column=20+ci).value = temp
              sheet.cell(row= k+1+ri, column=21+ci).value = temp1
              T=T+1
      
      else:
        #print("else")
        T=T+1

######################################################################################################sorting learning 23 24 ..... 26 27
    for T in range(MAX-MIN+1):
       for R in range(((MAX-MIN+1)-1) - T):

          if sheet.cell(row= R+2+ri, column=23+ci).value > sheet.cell(row= R+1+ri, column=23+ci).value:
            temp = sheet.cell(row= R+2+ri, column=23+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=24+ci).value
            sheet.cell(row= R+2+ri, column=23+ci).value = sheet.cell(row= R+1+ri, column=23+ci).value
            sheet.cell(row= R+2+ri, column=24+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value
            sheet.cell(row= R+1+ri, column=23+ci).value = temp
            sheet.cell(row= R+1+ri, column=24+ci).value = temp1  


            temp = sheet.cell(row= R+2+ri, column=26+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=27+ci).value
            sheet.cell(row= R+2+ri, column=26+ci).value = sheet.cell(row= R+1+ri, column=26+ci).value
            sheet.cell(row= R+2+ri, column=27+ci).value = sheet.cell(row= R+1+ri, column=27+ci).value
            sheet.cell(row= R+1+ri, column=26+ci).value = temp
            sheet.cell(row= R+1+ri, column=27+ci).value = temp1  




#sorting
    for T in range(MAX-MIN+1):
       for R in range(((MAX-MIN+1)-1) - T):

          if sheet.cell(row= R+2+ri, column=24+ci).value > sheet.cell(row= R+1+ri, column=24+ci).value:
            temp = sheet.cell(row= R+2+ri, column=23+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=24+ci).value
            sheet.cell(row= R+2+ri, column=23+ci).value = sheet.cell(row= R+1+ri, column=23+ci).value
            sheet.cell(row= R+2+ri, column=24+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value
            sheet.cell(row= R+1+ri, column=23+ci).value = temp
            sheet.cell(row= R+1+ri, column=24+ci).value = temp1        



            temp = sheet.cell(row= R+2+ri, column=26+ci).value
            temp1 = sheet.cell(row= R+2+ri, column=27+ci).value
            sheet.cell(row= R+2+ri, column=26+ci).value = sheet.cell(row= R+1+ri, column=26+ci).value
            sheet.cell(row= R+2+ri, column=27+ci).value = sheet.cell(row= R+1+ri, column=27+ci).value
            sheet.cell(row= R+1+ri, column=26+ci).value = temp
            sheet.cell(row= R+1+ri, column=27+ci).value = temp1  

#############################################################################################################Final Sorting learning 23 24 ..... 26 27
    pointer= sheet.cell(row= 1+ri, column=23+ci).value
    er=0
    T=0
    while(T<MAX-MIN+1):
      
      #print("er"+str(er))
      #print("T"+str(T))

      if sheet.cell(row= T+1+ri, column=24+ci).value == 0:
        er=er+1
        for k in range(MAX-MIN+1):
          if pointer + er<= MAX :
            if (sheet.cell(row= k+1+ri, column=24+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=23+ci).value == pointer +er ):
              temp = sheet.cell(row= T+1+ri, column=23+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=24+ci).value
              sheet.cell(row= T+1+ri, column=23+ci).value = sheet.cell(row= k+1+ri, column=23+ci).value
              sheet.cell(row= T+1+ri, column=24+ci).value = sheet.cell(row= k+1+ri, column=24+ci).value
              sheet.cell(row= k+1+ri, column=23+ci).value = temp
              sheet.cell(row= k+1+ri, column=24+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=26+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=27+ci).value
              sheet.cell(row= T+1+ri, column=26+ci).value = sheet.cell(row= k+1+ri, column=26+ci).value
              sheet.cell(row= T+1+ri, column=27+ci).value = sheet.cell(row= k+1+ri, column=27+ci).value
              sheet.cell(row= k+1+ri, column=26+ci).value = temp
              sheet.cell(row= k+1+ri, column=27+ci).value = temp1
              T=T+1

          if pointer - er>= MIN :
            if (sheet.cell(row= k+1+ri, column=24+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=23+ci).value == pointer -er ):
              temp = sheet.cell(row= T+1+ri, column=23+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=24+ci).value
              sheet.cell(row= T+1+ri, column=23+ci).value = sheet.cell(row= k+1+ri, column=23+ci).value
              sheet.cell(row= T+1+ri, column=24+ci).value = sheet.cell(row= k+1+ri, column=24+ci).value
              sheet.cell(row= k+1+ri, column=23+ci).value = temp
              sheet.cell(row= k+1+ri, column=24+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=26+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=27+ci).value
              sheet.cell(row= T+1+ri, column=26+ci).value = sheet.cell(row= k+1+ri, column=26+ci).value
              sheet.cell(row= T+1+ri, column=27+ci).value = sheet.cell(row= k+1+ri, column=27+ci).value
              sheet.cell(row= k+1+ri, column=26+ci).value = temp
              sheet.cell(row= k+1+ri, column=27+ci).value = temp1
              T=T+1
      
      else:
        #print("else")
        T=T+1


    #print("While End")



##########################################################################vazn
    StaticBit=SStaticBit
    DynamicBit=DDynamicBit
    ChangeBit= int((MAX-MIN+1)/DynamicBit)
    for R in range(MIN,MAX+1):
        if ((R+1)%ChangeBit)==0 and DynamicBit != 1 :
            DynamicBit= DynamicBit-1  


        if sheet.cell(row= R+1+ri, column=7+ci).value != 0:
            sheet.cell(row= R+1+ri, column=7+ci).value =sheet.cell(row= R+1+ri, column=7+ci).value + (StaticBit)
        else: 
            sheet.cell(row= R+1+ri, column=7+ci).value =sheet.cell(row= R+1+ri, column=7+ci).value + (DynamicBit)


##########################################################################vazn learning
    StaticBit=SStaticBit
    DynamicBit=DDynamicBit
    ChangeBit= int((MAX-MIN+1)/DynamicBit)
    for R in range(MIN,MAX+1):
        if ((R+1)%ChangeBit)==0 and DynamicBit != 1 :
            DynamicBit= DynamicBit-1  

        if sheet.cell(row= R+1+ri, column=24+ci).value != 0:
            sheet.cell(row= R+1+ri, column=24+ci).value =sheet.cell(row= R+1+ri, column=24+ci).value + (StaticBit)
        else: 
            sheet.cell(row= R+1+ri, column=24+ci).value =sheet.cell(row= R+1+ri, column=24+ci).value + (DynamicBit)




#########################################################################################################
#mohasebe
#copy
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=9+ci).value = str(sheet.cell(row= R+1+ri, column=6+ci).value)
       sheet.cell(row= R+1+ri, column=12+ci).value = str(sheet.cell(row= R+1+ri, column=6+ci).value)

       sheet.cell(row= R+1+ri, column=10+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
       sheet.cell(row= R+1+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value

       sheet.cell(row= R+1+ri, column=14+ci).value = '*'
       sheet.cell(row= R+1+ri, column=15+ci).value = '*'

############################################################################# Preparation learning 23 24 ....>> 29 30.. 32 33... 34 35
#copy
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=29+ci).value = str(sheet.cell(row= R+1+ri, column=23+ci).value)
       sheet.cell(row= R+1+ri, column=32+ci).value = str(sheet.cell(row= R+1+ri, column=23+ci).value)

       sheet.cell(row= R+1+ri, column=30+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value
       sheet.cell(row= R+1+ri, column=33+ci).value = sheet.cell(row= R+1+ri, column=24+ci).value

       sheet.cell(row= R+1+ri, column=34+ci).value = '*'
       sheet.cell(row= R+1+ri, column=35+ci).value = '*'
################################################################################
    for k in range(MAX-MIN):
       R=(MAX-MIN+1)-(k+1)

       STe=sheet.cell(row= R+1+ri, column=9+ci).value + sheet.cell(row= R+ri, column=9+ci).value
       VTe=sheet.cell(row= R+1+ri, column=10+ci).value + sheet.cell(row= R+ri, column=10+ci).value

       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=12+ci).value = STe
       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=13+ci).value = VTe
       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=14+ci).value = sheet.cell(row= R+1+ri, column=9+ci).value
       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=15+ci).value = sheet.cell(row= R+ri, column=9+ci).value

       E=R
       while E>=1:
        if E==1:
          B=E
          WM=MAX-MIN+1
          while B< WM:
            sheet.cell(row= WM+ri, column=9+ci).value=sheet.cell(row= WM+ri-1, column=9+ci).value
            sheet.cell(row= WM+ri, column=10+ci).value=sheet.cell(row= WM+ri-1, column=10+ci).value

            WM=WM-1

          sheet.cell(row= E+ri, column=9+ci).value = STe
          sheet.cell(row= E+ri, column=10+ci).value = VTe
          break
          
        else:
           if VTe<= (sheet.cell(row= E+ri-1, column=10+ci).value):
            B=E
            WM=MAX-MIN+1

            while B< WM:
              sheet.cell(row= WM+ri, column=9+ci).value=sheet.cell(row= WM+ri-1, column=9+ci).value
              sheet.cell(row= WM+ri, column=10+ci).value=sheet.cell(row= WM+ri-1, column=10+ci).value

              WM=WM-1

            sheet.cell(row= E+ri, column=9+ci).value = STe
            sheet.cell(row= E+ri, column=10+ci).value = VTe
            break

        E=E-1

    for R in range(MAX-MIN+1):
       RTe=sheet.cell(row= R+1+ri, column=12+ci).value
       coun=0
       End=MAX-MIN+1+MAX-MIN
       FSTe=""
       while RTe != sheet.cell(row= 1+ri, column=9+ci).value:
         Ind=MAX-MIN+1
         while Ind< End :
          if sheet.cell(row= Ind+1+ri, column=14+ci).value==RTe:
            FSTe=FSTe+"0"
            coun=coun+1
            RTe=sheet.cell(row= Ind+1+ri, column=12+ci).value
            break

          if sheet.cell(row= Ind+1+ri, column=15+ci).value==RTe:
            FSTe=FSTe+"1"
            coun=coun+1
            RTe=sheet.cell(row= Ind+1+ri, column=12+ci).value
            break
          Ind=Ind+1
       sheet.cell(row= R+1+ri, column=17+ci).value =  FSTe
       sheet.cell(row= R+1+ri, column=18+ci).value =  coun



################################################################################huffman learning  29 30.. 32 33... 34 35 .. 37 38
    for k in range(MAX-MIN):
       R=(MAX-MIN+1)-(k+1)

       STe=sheet.cell(row= R+1+ri, column=29+ci).value + sheet.cell(row= R+ri, column=29+ci).value
       VTe=sheet.cell(row= R+1+ri, column=30+ci).value + sheet.cell(row= R+ri, column=30+ci).value

       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=32+ci).value = STe
       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=33+ci).value = VTe
       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=34+ci).value = sheet.cell(row= R+1+ri, column=29+ci).value
       sheet.cell(row= k+1+ MAX-MIN+1+ri, column=35+ci).value = sheet.cell(row= R+ri, column=29+ci).value

       E=R
       while E>=1:
        if E==1:
          B=E
          WM=MAX-MIN+1
          while B< WM:
            sheet.cell(row= WM+ri, column=29+ci).value=sheet.cell(row= WM+ri-1, column=29+ci).value
            sheet.cell(row= WM+ri, column=30+ci).value=sheet.cell(row= WM+ri-1, column=30+ci).value

            WM=WM-1

          sheet.cell(row= E+ri, column=29+ci).value = STe
          sheet.cell(row= E+ri, column=30+ci).value = VTe
          break
          
        else:
           if VTe<= (sheet.cell(row= E+ri-1, column=30+ci).value):
            B=E
            WM=MAX-MIN+1

            while B< WM:
              sheet.cell(row= WM+ri, column=29+ci).value=sheet.cell(row= WM+ri-1, column=29+ci).value
              sheet.cell(row= WM+ri, column=30+ci).value=sheet.cell(row= WM+ri-1, column=30+ci).value

              WM=WM-1

            sheet.cell(row= E+ri, column=29+ci).value = STe
            sheet.cell(row= E+ri, column=30+ci).value = VTe
            break

        E=E-1

    for R in range(MAX-MIN+1):
       RTe=sheet.cell(row= R+1+ri, column=32+ci).value
       coun=0
       End=MAX-MIN+1+MAX-MIN
       FSTe=""
       while RTe != sheet.cell(row= 1+ri, column=29+ci).value:
         Ind=MAX-MIN+1
         while Ind< End :
          if sheet.cell(row= Ind+1+ri, column=34+ci).value==RTe:
            FSTe=FSTe+"0"
            coun=coun+1
            RTe=sheet.cell(row= Ind+1+ri, column=32+ci).value
            break

          if sheet.cell(row= Ind+1+ri, column=35+ci).value==RTe:
            FSTe=FSTe+"1"
            coun=coun+1
            RTe=sheet.cell(row= Ind+1+ri, column=32+ci).value
            break
          Ind=Ind+1
       sheet.cell(row= R+1+ri, column=37+ci).value =  FSTe
       sheet.cell(row= R+1+ri, column=38+ci).value =  coun







##########################################333
    x=0
    xl=0
    for R in range(MAX-MIN+1):
      sheet.cell(row= R+1+ri, column=19+ci).value = (sheet.cell(row= R+1+ri, column=18+ci).value)*(sheet.cell(row= R+1+ri, column=21+ci).value)
      x=x+sheet.cell(row= R+1+ri, column=19+ci).value


      sheet.cell(row= R+1+ri, column=39+ci).value = (sheet.cell(row= R+1+ri, column=38+ci).value)*(sheet.cell(row= R+1+ri, column=27+ci).value)
      xl=xl+sheet.cell(row= R+1+ri, column=39+ci).value


    p=p+x
    pl=pl+xl





           
#########################################################################################3


  
    print('\nSStaticBit: ')
    print(SStaticBit)
    print('\nDDynamicBit: ')
    print(DDynamicBit)



    print('\nX: ')
    print(p)

    print('\nXL: ')
    print(pl)

    print('\nXL/X: ')
    print(pl/p)

    sheet.cell(row= 1+ri, column=42+ci).value= "Normal"
    uu=NUM*(math.ceil(math.log(MAX-MIN+1,2)))
    sheet.cell(row= 2+ri, column=42+ci).value= uu


    sheet.cell(row= 1+ri, column=43+ci).value= "Cluster"
    sheet.cell(row= 2+ri, column=43+ci).value= p
    sheet.cell(row= 5+ri, column=43+ci).value= p/uu
    
    sheet.cell(row= 1+ri, column=44+ci).value= "Cluster_Learning"
    sheet.cell(row= 2+ri, column=44+ci).value= pl
    sheet.cell(row= 5+ri, column=44+ci).value= pl/uu

    sheet.cell(row= 1+ri, column=45+ci).value= "Cluster_Learning/Cluster"
    sheet.cell(row= 2+ri, column=45+ci).value= pl/p
    sheet.cell(row= 5+ri, column=45+ci).value= (pl/p)*100
    sheet.cell(row= 7+ri, column=45+ci).value= round((pl/p)*100)  


    c3 = arr.array('i') 
    c4 = arr.array('i') 
    c1 = arr.array('i') 
    c11 = arr.array('i') 
    for R in range(NUM):
    	c1.append(sheet.cell(row= R+1+ri, column=1+ci).value)

    for R in range(365):
    	c11.append(sheet.cell(row= R+1+ri, column=1+ci).value)    	

    for R in range(MAX-MIN+1):
       c3.append( sheet.cell(row= R+1+ri, column=3+ci).value)
       c4.append( sheet.cell(row= R+1+ri, column=4+ci).value)



    # plt.hist(c1,bins=math.ceil((MAX-MIN+1)/3),density=True)
    # plt.xlabel('Temperature (C)')
    # plt.ylabel('Frequency (%)')
    # plt.title('Frequency chart')

    # #plt.hist(c1,density=True)
    # plt.xlim([MIN, MAX])

    # plt.show()

    # plt.plot(c11)
    # plt.ylabel('Temperature (C)')
    # plt.xlabel('Day')
    # plt.title('Temperature changes over 365 days')
    # plt.show()

    # Ndata=96341
    # plt.bar(['Clustered','Normal'],[(x/(Ndata))*100,(Ndata/Ndata)*100])
    # plt.ylabel('Energy')
    # plt.xlabel('Data')
    # plt.title('Energy consumption with SDN')
    # plt.show()


    if (NMM/365)<=1:
      book.save('SampleTG_SDN_AnalyzedN_H_R_W_L_P____S'+str(SStaticBit)+'D'+str(DDynamicBit)+'_'+str(NMM)+'.xlsx')
    else:
     book.save('SampleTG_SDN_AnalyzedN_H365Z_R_W_L_P____S'+str(SStaticBit)+'D'+str(DDynamicBit)+'_'+str(int(NMM/365))+'.xlsx')






   # book.save('SampleTG_SDN_AnalyzedN_H.xlsx')
if __name__ == '__main__':
    main()