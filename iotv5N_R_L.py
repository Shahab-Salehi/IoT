import openpyxl
import time 
import math
import matplotlib.pyplot as plt
import array as arr

def main():
    #book = openpyxl.load_workbook('C:\Users\shahab\OneDrive\FA_paper\py\Sample.xlsx')
    book = openpyxl.load_workbook('temp_mean_105_days_actual_predictions.xlsx')
    #book = openpyxl.Workbook()
   # book.create_sheet('Sample')


    #sheet = book["Sample"]
    sheet = book.active



    NUM = 2612
    MAX=80
    MIN=-46
    ri=0
    ci=0
    NMM=105

#ijad magadir beyn max v min , v barbar sefr garar dadan un ha
    for R in range(MIN,MAX+1):
       sheet.cell(row= R-MIN+1+ri, column=3+ci).value = R
       sheet.cell(row= R-MIN+1+ri, column=4+ci).value = 0

#shomaresh magadir
    for R in range(NMM):
          sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value = sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value+1

#copy
    for R in range(MAX-MIN+1):
      sheet.cell(row= R+1+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
    for R in range(MAX-MIN+1):
      sheet.cell(row= R+1+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value

#copy2
    for R in range(MAX-MIN+1):
      sheet.cell(row= R+1+ri, column=12+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
    for R in range(MAX-MIN+1):
      sheet.cell(row= R+1+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value




# #sorting
#     for T in range(MAX-MIN+1):
#       for R in range(((MAX-MIN+1)-1) - T):
#         if sheet.cell(row= R+2+ri, column=6+ci).value > sheet.cell(row= R+1+ri, column=6+ci).value:
#           temp = sheet.cell(row= R+2+ri, column=6+ci).value
#           temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
#           sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
#           sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
#           sheet.cell(row= R+1+ri, column=6+ci).value = temp
#           sheet.cell(row= R+1+ri, column=7+ci).value = temp1 

#     for T in range(MAX-MIN+1):
#       for R in range(((MAX-MIN+1)-1) - T):
#         if sheet.cell(row= R+2+ri, column=7+ci).value > sheet.cell(row= R+1+ri, column=7+ci).value:
#           temp = sheet.cell(row= R+2+ri, column=6+ci).value
#           temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
#           sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
#           sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
#           sheet.cell(row= R+1+ri, column=6+ci).value = temp
#           sheet.cell(row= R+1+ri, column=7+ci).value = temp1

#ijad magadir beyn max v min , v barbar sefr garar dadan un ha
    for R in range(MIN,MAX+1):
      sheet.cell(row= R-MIN+1+ri, column=3+ci).value = R
      sheet.cell(row= R-MIN+1+ri, column=4+ci).value = 0
      sheet.cell(row= R-MIN+1+ri, column=15+ci).value = R
      sheet.cell(row= R-MIN+1+ri, column=16+ci).value = 0






###########################################################################shomaresh magadir
    p=0
    pl=0
    for R in range(NMM,NUM):
          sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value = sheet.cell(row= (sheet.cell(row= R+1+ri, column=1+ci).value)-MIN+1+ri, column=4+ci).value+1
          sheet.cell(row= (sheet.cell(row= R+1+ri, column=2+ci).value)-MIN+1+ri, column=16+ci).value = sheet.cell(row= (sheet.cell(row= R+1+ri, column=2+ci).value)-MIN+1+ri, column=16+ci).value+1
          if ((R+1)%NMM)==0:
#############################################################################copy
                print(round((R/NUM)*100), end="\r")
                for R in range(MAX-MIN+1):
                    sheet.cell(row= R+1+ri, column=12+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
                for R in range(MAX-MIN+1):
                    sheet.cell(row= R+1+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value
#############################################################################copy learning

                for R in range(MAX-MIN+1):
                    sheet.cell(row= R+1+ri, column=18+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
                for R in range(MAX-MIN+1):
                    sheet.cell(row= R+1+ri, column=19+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value



###########################################################################sorting
                for T in range(MAX-MIN+1):
                  for R in range(((MAX-MIN+1)-1) - T):
                    if sheet.cell(row= R+2+ri, column=6+ci).value > sheet.cell(row= R+1+ri, column=6+ci).value:
                      temp = sheet.cell(row= R+2+ri, column=6+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
                      sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
                      sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
                      sheet.cell(row= R+1+ri, column=6+ci).value = temp
                      sheet.cell(row= R+1+ri, column=7+ci).value = temp1

                      temp = sheet.cell(row= R+2+ri, column=12+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=13+ci).value
                      sheet.cell(row= R+2+ri, column=12+ci).value = sheet.cell(row= R+1+ri, column=12+ci).value
                      sheet.cell(row= R+2+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=13+ci).value
                      sheet.cell(row= R+1+ri, column=12+ci).value = temp
                      sheet.cell(row= R+1+ri, column=13+ci).value = temp1 

                for T in range(MAX-MIN+1):
                  for R in range(((MAX-MIN+1)-1) - T):
                    if sheet.cell(row= R+2+ri, column=7+ci).value > sheet.cell(row= R+1+ri, column=7+ci).value:
                      temp = sheet.cell(row= R+2+ri, column=6+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
                      sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
                      sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
                      sheet.cell(row= R+1+ri, column=6+ci).value = temp
                      sheet.cell(row= R+1+ri, column=7+ci).value = temp1

                      temp = sheet.cell(row= R+2+ri, column=12+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=13+ci).value
                      sheet.cell(row= R+2+ri, column=12+ci).value = sheet.cell(row= R+1+ri, column=12+ci).value
                      sheet.cell(row= R+2+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=13+ci).value
                      sheet.cell(row= R+1+ri, column=12+ci).value = temp
                      sheet.cell(row= R+1+ri, column=13+ci).value = temp1


###########################################################################Final Sorting
                pointer= sheet.cell(row= 1+ri, column=6+ci).value
                er=0
                T=0
                while(T<MAX-MIN+1):
                  
                  #print("er"+str(er))
                  #print("T"+str(T))

                  if sheet.cell(row= T+1+ri, column=7+ci).value == 0:
                    er=er+1
                    for k in range(MAX-MIN+1):
                      if pointer + er<= MAX :
                        if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer +er ):
                          temp = sheet.cell(row= T+1+ri, column=6+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
                          sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
                          sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
                          sheet.cell(row= k+1+ri, column=6+ci).value = temp
                          sheet.cell(row= k+1+ri, column=7+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=12+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=13+ci).value
                          sheet.cell(row= T+1+ri, column=12+ci).value = sheet.cell(row= k+1+ri, column=12+ci).value
                          sheet.cell(row= T+1+ri, column=13+ci).value = sheet.cell(row= k+1+ri, column=13+ci).value
                          sheet.cell(row= k+1+ri, column=12+ci).value = temp
                          sheet.cell(row= k+1+ri, column=13+ci).value = temp1
                          T=T+1

                      if pointer - er>= MIN :
                        if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer -er ):
                          temp = sheet.cell(row= T+1+ri, column=6+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
                          sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
                          sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
                          sheet.cell(row= k+1+ri, column=6+ci).value = temp
                          sheet.cell(row= k+1+ri, column=7+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=12+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=13+ci).value
                          sheet.cell(row= T+1+ri, column=12+ci).value = sheet.cell(row= k+1+ri, column=12+ci).value
                          sheet.cell(row= T+1+ri, column=13+ci).value = sheet.cell(row= k+1+ri, column=13+ci).value
                          sheet.cell(row= k+1+ri, column=12+ci).value = temp
                          sheet.cell(row= k+1+ri, column=13+ci).value = temp1
                          T=T+1
                  
                  else:
                    #print("else")
                    T=T+1


                #print("While End")


###########################################################################sorting learning 15 16.....18 19
                for T in range(MAX-MIN+1):
                  for R in range(((MAX-MIN+1)-1) - T):
                    if sheet.cell(row= R+2+ri, column=15+ci).value > sheet.cell(row= R+1+ri, column=15+ci).value:
                      temp = sheet.cell(row= R+2+ri, column=15+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=16+ci).value
                      sheet.cell(row= R+2+ri, column=15+ci).value = sheet.cell(row= R+1+ri, column=15+ci).value
                      sheet.cell(row= R+2+ri, column=16+ci).value = sheet.cell(row= R+1+ri, column=16+ci).value
                      sheet.cell(row= R+1+ri, column=15+ci).value = temp
                      sheet.cell(row= R+1+ri, column=16+ci).value = temp1

                      temp = sheet.cell(row= R+2+ri, column=18+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=19+ci).value
                      sheet.cell(row= R+2+ri, column=18+ci).value = sheet.cell(row= R+1+ri, column=18+ci).value
                      sheet.cell(row= R+2+ri, column=19+ci).value = sheet.cell(row= R+1+ri, column=19+ci).value
                      sheet.cell(row= R+1+ri, column=18+ci).value = temp
                      sheet.cell(row= R+1+ri, column=19+ci).value = temp1 

                for T in range(MAX-MIN+1):
                  for R in range(((MAX-MIN+1)-1) - T):
                    if sheet.cell(row= R+2+ri, column=16+ci).value > sheet.cell(row= R+1+ri, column=16+ci).value:
                      temp = sheet.cell(row= R+2+ri, column=15+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=16+ci).value
                      sheet.cell(row= R+2+ri, column=15+ci).value = sheet.cell(row= R+1+ri, column=15+ci).value
                      sheet.cell(row= R+2+ri, column=16+ci).value = sheet.cell(row= R+1+ri, column=16+ci).value
                      sheet.cell(row= R+1+ri, column=15+ci).value = temp
                      sheet.cell(row= R+1+ri, column=16+ci).value = temp1

                      temp = sheet.cell(row= R+2+ri, column=18+ci).value
                      temp1 = sheet.cell(row= R+2+ri, column=19+ci).value
                      sheet.cell(row= R+2+ri, column=18+ci).value = sheet.cell(row= R+1+ri, column=18+ci).value
                      sheet.cell(row= R+2+ri, column=19+ci).value = sheet.cell(row= R+1+ri, column=19+ci).value
                      sheet.cell(row= R+1+ri, column=18+ci).value = temp
                      sheet.cell(row= R+1+ri, column=19+ci).value = temp1


###########################################################################Final Sorting learning 15 16.....18 19
                pointer= sheet.cell(row= 1+ri, column=15+ci).value
                er=0
                T=0
                while(T<MAX-MIN+1):
                  
                  #print("er"+str(er))
                  #print("T"+str(T))

                  if sheet.cell(row= T+1+ri, column=16+ci).value == 0:
                    er=er+1
                    for k in range(MAX-MIN+1):
                      if pointer + er<= MAX :
                        if (sheet.cell(row= k+1+ri, column=16+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=15+ci).value == pointer +er ):
                          temp = sheet.cell(row= T+1+ri, column=15+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=16+ci).value
                          sheet.cell(row= T+1+ri, column=15+ci).value = sheet.cell(row= k+1+ri, column=15+ci).value
                          sheet.cell(row= T+1+ri, column=16+ci).value = sheet.cell(row= k+1+ri, column=16+ci).value
                          sheet.cell(row= k+1+ri, column=15+ci).value = temp
                          sheet.cell(row= k+1+ri, column=16+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=18+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=19+ci).value
                          sheet.cell(row= T+1+ri, column=18+ci).value = sheet.cell(row= k+1+ri, column=18+ci).value
                          sheet.cell(row= T+1+ri, column=19+ci).value = sheet.cell(row= k+1+ri, column=19+ci).value
                          sheet.cell(row= k+1+ri, column=18+ci).value = temp
                          sheet.cell(row= k+1+ri, column=19+ci).value = temp1
                          T=T+1

                      if pointer - er>= MIN :
                        if (sheet.cell(row= k+1+ri, column=16+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=15+ci).value == pointer -er ):
                          temp = sheet.cell(row= T+1+ri, column=15+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=16+ci).value
                          sheet.cell(row= T+1+ri, column=15+ci).value = sheet.cell(row= k+1+ri, column=15+ci).value
                          sheet.cell(row= T+1+ri, column=16+ci).value = sheet.cell(row= k+1+ri, column=16+ci).value
                          sheet.cell(row= k+1+ri, column=15+ci).value = temp
                          sheet.cell(row= k+1+ri, column=16+ci).value = temp1

                          temp = sheet.cell(row= T+1+ri, column=18+ci).value
                          temp1 = sheet.cell(row= T+1+ri, column=19+ci).value
                          sheet.cell(row= T+1+ri, column=18+ci).value = sheet.cell(row= k+1+ri, column=18+ci).value
                          sheet.cell(row= T+1+ri, column=19+ci).value = sheet.cell(row= k+1+ri, column=19+ci).value
                          sheet.cell(row= k+1+ri, column=18+ci).value = temp
                          sheet.cell(row= k+1+ri, column=19+ci).value = temp1
                          T=T+1
                  
                  else:
                    #print("else")
                    T=T+1


                #print("While End")

#################################################################coding and mohasebe
                x=0
                xl=0
                for R in range(MAX-MIN+1):
                  sheet.cell(row= R+1+ri, column=8+ci).value = math.floor(math.log((R+1)*8,2))
                  sheet.cell(row= R+1+ri, column=9+ci).value=(sheet.cell(row= R+1+ri, column=13+ci).value)*(sheet.cell(row= R+1+ri, column=8+ci).value)
                  x=x + sheet.cell(row= R+1+ri, column=9+ci).value
                  sheet.cell(row= R+1+ri, column=10+ci).value=x

                  sheet.cell(row= R+1+ri, column=20+ci).value = math.floor(math.log((R+1)*8,2))
                  sheet.cell(row= R+1+ri, column=21+ci).value=(sheet.cell(row= R+1+ri, column=19+ci).value)*(sheet.cell(row= R+1+ri, column=20+ci).value)
                  xl=xl + sheet.cell(row= R+1+ri, column=21+ci).value
                  sheet.cell(row= R+1+ri, column=22+ci).value=xl


                p=p+x
                pl=pl+xl
#############################################################################copy
                for R in range(MAX-MIN+1):
                  sheet.cell(row= R+1+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
                for R in range(MAX-MIN+1):
                  sheet.cell(row= R+1+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value

                # for T in range(MAX-MIN+1):
                #   for R in range(((MAX-MIN+1)-1) - T):
                #     if sheet.cell(row= R+2+ri, column=6+ci).value > sheet.cell(row= R+1+ri, column=6+ci).value:
                #       temp = sheet.cell(row= R+2+ri, column=6+ci).value
                #       temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
                #       sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
                #       sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
                #       sheet.cell(row= R+1+ri, column=6+ci).value = temp
                #       sheet.cell(row= R+1+ri, column=7+ci).value = temp1 

                # for T in range(MAX-MIN+1):
                #   for R in range(((MAX-MIN+1)-1) - T):
                #     if sheet.cell(row= R+2+ri, column=7+ci).value > sheet.cell(row= R+1+ri, column=7+ci).value:
                #       temp = sheet.cell(row= R+2+ri, column=6+ci).value
                #       temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
                #       sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
                #       sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
                #       sheet.cell(row= R+1+ri, column=6+ci).value = temp
                #       sheet.cell(row= R+1+ri, column=7+ci).value = temp1

                for R in range(MIN,MAX+1):
                  sheet.cell(row= R-MIN+1+ri, column=3+ci).value = R
                  sheet.cell(row= R-MIN+1+ri, column=4+ci).value = 0

                for R in range(MIN,MAX+1):
                  sheet.cell(row= R-MIN+1+ri, column=15+ci).value = R
                  sheet.cell(row= R-MIN+1+ri, column=16+ci).value = 0






    for R in range(MAX-MIN+1):
        sheet.cell(row= R+1+ri, column=12+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
    for R in range(MAX-MIN+1):
        sheet.cell(row= R+1+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value

    for R in range(MAX-MIN+1):
        sheet.cell(row= R+1+ri, column=18+ci).value = sheet.cell(row= R+1+ri, column=3+ci).value
    for R in range(MAX-MIN+1):
        sheet.cell(row= R+1+ri, column=19+ci).value = sheet.cell(row= R+1+ri, column=4+ci).value



    for T in range(MAX-MIN+1):
      for R in range(((MAX-MIN+1)-1) - T):
        if sheet.cell(row= R+2+ri, column=6+ci).value > sheet.cell(row= R+1+ri, column=6+ci).value:
          temp = sheet.cell(row= R+2+ri, column=6+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
          sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
          sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
          sheet.cell(row= R+1+ri, column=6+ci).value = temp
          sheet.cell(row= R+1+ri, column=7+ci).value = temp1

          temp = sheet.cell(row= R+2+ri, column=12+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=13+ci).value
          sheet.cell(row= R+2+ri, column=12+ci).value = sheet.cell(row= R+1+ri, column=12+ci).value
          sheet.cell(row= R+2+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=13+ci).value
          sheet.cell(row= R+1+ri, column=12+ci).value = temp
          sheet.cell(row= R+1+ri, column=13+ci).value = temp1 

    for T in range(MAX-MIN+1):
      for R in range(((MAX-MIN+1)-1) - T):
        if sheet.cell(row= R+2+ri, column=7+ci).value > sheet.cell(row= R+1+ri, column=7+ci).value:
          temp = sheet.cell(row= R+2+ri, column=6+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=7+ci).value
          sheet.cell(row= R+2+ri, column=6+ci).value = sheet.cell(row= R+1+ri, column=6+ci).value
          sheet.cell(row= R+2+ri, column=7+ci).value = sheet.cell(row= R+1+ri, column=7+ci).value
          sheet.cell(row= R+1+ri, column=6+ci).value = temp
          sheet.cell(row= R+1+ri, column=7+ci).value = temp1

          temp = sheet.cell(row= R+2+ri, column=12+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=13+ci).value
          sheet.cell(row= R+2+ri, column=12+ci).value = sheet.cell(row= R+1+ri, column=12+ci).value
          sheet.cell(row= R+2+ri, column=13+ci).value = sheet.cell(row= R+1+ri, column=13+ci).value
          sheet.cell(row= R+1+ri, column=12+ci).value = temp
          sheet.cell(row= R+1+ri, column=13+ci).value = temp1








    pointer= sheet.cell(row= 1+ri, column=6+ci).value
    er=0
    T=0
    while(T<MAX-MIN+1):
      if sheet.cell(row= T+1+ri, column=7+ci).value == 0:
        er=er+1
        for k in range(MAX-MIN+1):
          if pointer + er<= MAX :
            if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer +er ):
              temp = sheet.cell(row= T+1+ri, column=6+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
              sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
              sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
              sheet.cell(row= k+1+ri, column=6+ci).value = temp
              sheet.cell(row= k+1+ri, column=7+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=12+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=13+ci).value
              sheet.cell(row= T+1+ri, column=12+ci).value = sheet.cell(row= k+1+ri, column=12+ci).value
              sheet.cell(row= T+1+ri, column=13+ci).value = sheet.cell(row= k+1+ri, column=13+ci).value
              sheet.cell(row= k+1+ri, column=12+ci).value = temp
              sheet.cell(row= k+1+ri, column=13+ci).value = temp1
              T=T+1

          if pointer - er>= MIN :
            if (sheet.cell(row= k+1+ri, column=7+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=6+ci).value == pointer -er ):
              temp = sheet.cell(row= T+1+ri, column=6+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=7+ci).value
              sheet.cell(row= T+1+ri, column=6+ci).value = sheet.cell(row= k+1+ri, column=6+ci).value
              sheet.cell(row= T+1+ri, column=7+ci).value = sheet.cell(row= k+1+ri, column=7+ci).value
              sheet.cell(row= k+1+ri, column=6+ci).value = temp
              sheet.cell(row= k+1+ri, column=7+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=12+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=13+ci).value
              sheet.cell(row= T+1+ri, column=12+ci).value = sheet.cell(row= k+1+ri, column=12+ci).value
              sheet.cell(row= T+1+ri, column=13+ci).value = sheet.cell(row= k+1+ri, column=13+ci).value
              sheet.cell(row= k+1+ri, column=12+ci).value = temp
              sheet.cell(row= k+1+ri, column=13+ci).value = temp1
              T=T+1
      
      else:
        T=T+1
##################################################################################################################################learning


##############sorting learning 15 16.....18 19
    for T in range(MAX-MIN+1):
      for R in range(((MAX-MIN+1)-1) - T):
        if sheet.cell(row= R+2+ri, column=15+ci).value > sheet.cell(row= R+1+ri, column=15+ci).value:
          temp = sheet.cell(row= R+2+ri, column=15+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=16+ci).value
          sheet.cell(row= R+2+ri, column=15+ci).value = sheet.cell(row= R+1+ri, column=15+ci).value
          sheet.cell(row= R+2+ri, column=16+ci).value = sheet.cell(row= R+1+ri, column=16+ci).value
          sheet.cell(row= R+1+ri, column=15+ci).value = temp
          sheet.cell(row= R+1+ri, column=16+ci).value = temp1

          temp = sheet.cell(row= R+2+ri, column=18+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=19+ci).value
          sheet.cell(row= R+2+ri, column=18+ci).value = sheet.cell(row= R+1+ri, column=18+ci).value
          sheet.cell(row= R+2+ri, column=19+ci).value = sheet.cell(row= R+1+ri, column=19+ci).value
          sheet.cell(row= R+1+ri, column=18+ci).value = temp
          sheet.cell(row= R+1+ri, column=19+ci).value = temp1 

    for T in range(MAX-MIN+1):
      for R in range(((MAX-MIN+1)-1) - T):
        if sheet.cell(row= R+2+ri, column=16+ci).value > sheet.cell(row= R+1+ri, column=16+ci).value:
          temp = sheet.cell(row= R+2+ri, column=15+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=16+ci).value
          sheet.cell(row= R+2+ri, column=15+ci).value = sheet.cell(row= R+1+ri, column=15+ci).value
          sheet.cell(row= R+2+ri, column=16+ci).value = sheet.cell(row= R+1+ri, column=16+ci).value
          sheet.cell(row= R+1+ri, column=15+ci).value = temp
          sheet.cell(row= R+1+ri, column=16+ci).value = temp1

          temp = sheet.cell(row= R+2+ri, column=18+ci).value
          temp1 = sheet.cell(row= R+2+ri, column=19+ci).value
          sheet.cell(row= R+2+ri, column=18+ci).value = sheet.cell(row= R+1+ri, column=18+ci).value
          sheet.cell(row= R+2+ri, column=19+ci).value = sheet.cell(row= R+1+ri, column=19+ci).value
          sheet.cell(row= R+1+ri, column=18+ci).value = temp
          sheet.cell(row= R+1+ri, column=19+ci).value = temp1


###############Final Sorting learning 15 16.....18 19
    pointer= sheet.cell(row= 1+ri, column=15+ci).value
    er=0
    T=0
    while(T<MAX-MIN+1):
      
      #print("er"+str(er))
      #print("T"+str(T))

      if sheet.cell(row= T+1+ri, column=16+ci).value == 0:
        er=er+1
        for k in range(MAX-MIN+1):
          if pointer + er<= MAX :
            if (sheet.cell(row= k+1+ri, column=16+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=15+ci).value == pointer +er ):
              temp = sheet.cell(row= T+1+ri, column=15+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=16+ci).value
              sheet.cell(row= T+1+ri, column=15+ci).value = sheet.cell(row= k+1+ri, column=15+ci).value
              sheet.cell(row= T+1+ri, column=16+ci).value = sheet.cell(row= k+1+ri, column=16+ci).value
              sheet.cell(row= k+1+ri, column=15+ci).value = temp
              sheet.cell(row= k+1+ri, column=16+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=18+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=19+ci).value
              sheet.cell(row= T+1+ri, column=18+ci).value = sheet.cell(row= k+1+ri, column=18+ci).value
              sheet.cell(row= T+1+ri, column=19+ci).value = sheet.cell(row= k+1+ri, column=19+ci).value
              sheet.cell(row= k+1+ri, column=18+ci).value = temp
              sheet.cell(row= k+1+ri, column=19+ci).value = temp1
              T=T+1

          if pointer - er>= MIN :
            if (sheet.cell(row= k+1+ri, column=16+ci).value == 0) and ( sheet.cell(row= k+1+ri, column=15+ci).value == pointer -er ):
              temp = sheet.cell(row= T+1+ri, column=15+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=16+ci).value
              sheet.cell(row= T+1+ri, column=15+ci).value = sheet.cell(row= k+1+ri, column=15+ci).value
              sheet.cell(row= T+1+ri, column=16+ci).value = sheet.cell(row= k+1+ri, column=16+ci).value
              sheet.cell(row= k+1+ri, column=15+ci).value = temp
              sheet.cell(row= k+1+ri, column=16+ci).value = temp1

              temp = sheet.cell(row= T+1+ri, column=18+ci).value
              temp1 = sheet.cell(row= T+1+ri, column=19+ci).value
              sheet.cell(row= T+1+ri, column=18+ci).value = sheet.cell(row= k+1+ri, column=18+ci).value
              sheet.cell(row= T+1+ri, column=19+ci).value = sheet.cell(row= k+1+ri, column=19+ci).value
              sheet.cell(row= k+1+ri, column=18+ci).value = temp
              sheet.cell(row= k+1+ri, column=19+ci).value = temp1
              T=T+1
      
      else:
        #print("else")
        T=T+1


    #print("While End")




###################################################################################################################################
#mohasebe
    x=0
    xl=0
    for R in range(MAX-MIN+1):
       sheet.cell(row= R+1+ri, column=8+ci).value = math.floor(math.log((R+1)*8,2))
       sheet.cell(row= R+1+ri, column=9+ci).value=(sheet.cell(row= R+1+ri, column=13+ci).value)*(sheet.cell(row= R+1+ri, column=8+ci).value)
       x=x + sheet.cell(row= R+1+ri, column=9+ci).value
       sheet.cell(row= R+1+ri, column=10+ci).value=x

       sheet.cell(row= R+1+ri, column=20+ci).value = math.floor(math.log((R+1)*8,2))
       sheet.cell(row= R+1+ri, column=21+ci).value=(sheet.cell(row= R+1+ri, column=19+ci).value)*(sheet.cell(row= R+1+ri, column=20+ci).value)
       xl=xl + sheet.cell(row= R+1+ri, column=21+ci).value
       sheet.cell(row= R+1+ri, column=22+ci).value=xl

    p=p+x
    pl=pl+xl

    print('\nX: ')
    print(p)

    print('\nXL: ')
    print(pl)

    print('\nXL/X: ')
    print(pl/p)

    sheet.cell(row= 1+ri, column=27+ci).value= "Normal"
    uu=NUM*(math.ceil(math.log(MAX-MIN+1,2)))
    sheet.cell(row= 2+ri, column=27+ci).value= uu


    sheet.cell(row= 1+ri, column=28+ci).value= "Cluster"
    sheet.cell(row= 2+ri, column=28+ci).value= p
    sheet.cell(row= 5+ri, column=28+ci).value= p/uu
    
    sheet.cell(row= 1+ri, column=29+ci).value= "Cluster_Learning"
    sheet.cell(row= 2+ri, column=29+ci).value= pl
    sheet.cell(row= 5+ri, column=29+ci).value= pl/uu

    sheet.cell(row= 1+ri, column=30+ci).value= "Cluster_Learning/Cluster"
    sheet.cell(row= 2+ri, column=30+ci).value= pl/p
    sheet.cell(row= 5+ri, column=30+ci).value= (pl/p)*100
    sheet.cell(row= 7+ri, column=30+ci).value= round((pl/p)*100)


    c3 = arr.array('i') 
    c4 = arr.array('i') 
    c1 = arr.array('i') 
    c11 = arr.array('i') 
    for R in range(NUM):
    	c1.append(sheet.cell(row= R+1+ri, column=1+ci).value)

    for R in range(365):
    	c11.append(sheet.cell(row= R+1+ri, column=1+ci).value)    	

    for R in range(MAX-MIN+1):
       c3.append( sheet.cell(row= R+1+ri, column=3+ci).value)
       c4.append( sheet.cell(row= R+1+ri, column=4+ci).value)



  #  plt.hist(c1,bins=math.ceil((MAX-MIN+1)/3),density=True)
   # plt.xlabel('Temperature (C)')
   # plt.ylabel('Frequency (%)')
   # plt.title('Frequency chart')

    #plt.hist(c1,density=True)
   # plt.xlim([MIN, MAX])

   # plt.show()

   # plt.plot(c11)
   # plt.ylabel('Temperature (C)')
   # plt.xlabel('Day')
   # plt.title('Temperature changes over 365 days')
   # plt.show()

   # Ndata=96341
   # plt.bar(['Clustered','Normal'],[(x/(Ndata))*100,(Ndata/Ndata)*100])
   # plt.ylabel('Energy')
   # plt.xlabel('Data')
   # plt.title('Energy consumption with SDN')
   # plt.show()




    if (NMM/365)<=1:
      book.save('SampleTG_SDN_AnalyzedN_R_L'+str(NMM)+'.xlsx')
    else:
     book.save('SampleTG_SDN_AnalyzedN365Z_R_L'+str(int(NMM/365))+'.xlsx')




    #book.save('SampleTG_SDN_AnalyzedN.xlsx')
if __name__ == '__main__':
    main()
